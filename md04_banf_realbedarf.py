#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Berechnet den bis zu einem Stichtag tatsaechlich benoetigten Anteil einer BANF aus einem SAP-MD04-Excel-Export.
python md04_banf_realbedarf.py EXPORT.xlsx --stichtag 31.12.2026 --banf
"""


from __future__ import annotations
import argparse
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

BANF_TYPES = {"BS-ANF", "BANF", "PURRQS"}
DEMAND_TYPES = {"AR-RES", "SEKBED", "KUND-B", "UML-RES", "AB-RES"}
STOCK_TYPES = {"BSTAND"}
SAFETY_TYPES = {"SHBEST"}

def norm(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).upper()


def parse_date(text: str) -> date:
    text = text.strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError("Datum ungueltig. Beispiele: 31.12.2026 oder 2026-12-31")


def cell_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        from openpyxl.utils.datetime import from_excel
        return from_excel(value).date()
    return parse_date(str(value))


def number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    return float(s)


@dataclass
class Md04Row:
    row_no: int
    plan_date: date
    element: str
    reference: str
    qty: float
    available: float | None


def find_column(headers: list[Any], candidates: Iterable[str]) -> int:
    normalized = [norm(x) for x in headers]
    for candidate in candidates:
        c = norm(candidate)
        if c in normalized:
            return normalized.index(c)
    raise KeyError(f"Spalte nicht gefunden: {', '.join(candidates)}; vorhanden: {headers}")


def read_md04(path: Path, sheet_name: str | None = None) -> list[Md04Row]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))

    i_date = find_column(headers, ["Plantermine", "Plantermin"])
    i_elem = find_column(headers, ["Dispoel.", "Dispoelement", "MRP-Element"])
    i_ref = find_column(headers, ["Daten zum DE", "MRP-Elementdaten"])
    i_qty = find_column(headers, ["Zugang/Bedarf", "Zugang / Bedarf"])
    try:
        i_avail = find_column(headers, ["Verfügb. Menge", "Verfuegbare Menge"])
    except KeyError:
        i_avail = -1

    result: list[Md04Row] = []
    for row_no, row in enumerate(rows, start=2):
        if not row[i_date] or not row[i_elem]:
            continue
        try:
            d = cell_date(row[i_date])
            q = number(row[i_qty])
        except (ValueError, TypeError):
            continue
        av = None
        if i_avail >= 0 and row[i_avail] not in (None, ""):
            try:
                av = number(row[i_avail])
            except ValueError:
                pass
        result.append(Md04Row(row_no, d, norm(row[i_elem]), str(row[i_ref] or "").strip(), q, av))
    return result


def choose_banf(rows: list[Md04Row], cutoff: date, banf_ref: str | None) -> Md04Row:
    candidates = [r for r in rows if r.element in BANF_TYPES]
    if banf_ref:
        wanted = norm(banf_ref)
        candidates = [r for r in candidates if norm(r.reference) == wanted]
    else:
        candidates = [r for r in candidates if r.plan_date <= cutoff]
    if not candidates:
        raise ValueError("Keine passende BANF gefunden.")
    return max(candidates, key=lambda r: (r.plan_date, r.row_no))


def calculate(rows: list[Md04Row], cutoff: date, target: Md04Row) -> dict[str, float]:
    # 口径二：只计算实际需求，允许使用安全库存。
    # 因此 ShBest 仅作为信息显示，不作为真实消耗参与净需求计算。
    stock_rows = [r for r in rows if r.element in STOCK_TYPES and r.row_no <= target.row_no]
    if not stock_rows:
        raise ValueError("Kein BStand/Anfangsbestand vor der BANF gefunden.")
    stock_row = max(stock_rows, key=lambda r: r.row_no)
    opening_stock = stock_row.available if stock_row.available is not None else stock_row.qty

    # MD04 可能包含早于当前日期、但仍显示在 BStand 后面的未结需求。
    # 所以按行位置从 BStand 之后开始，并按截止日期筛选。
    movements = [r for r in rows if r.row_no > stock_row.row_no and r.plan_date <= cutoff]

    safety_stock = -sum(
        r.qty for r in movements if r.element in SAFETY_TYPES and r.qty < 0
    )
    real_demand = -sum(
        r.qty for r in movements
        if r.qty < 0 and r.element not in SAFETY_TYPES
    )
    other_receipts = sum(
        r.qty for r in movements
        if r.row_no != target.row_no and r.qty > 0
    )

    # 真实净缺口 = 实际需求 - 初始库存 - 其他收入。
    # 目标 BANF 自身不算作“其他收入”，否则无法判断它真正需要多少。
    net_real_requirement = max(0.0, real_demand - opening_stock - other_receipts)
    usable_target = max(0.0, target.qty)
    real_needed_from_banf = min(usable_target, net_real_requirement)

    return {
        "balance_without_target": opening_stock + other_receipts - real_demand,
        "required_from_target": net_real_requirement,
        "real_needed": real_needed_from_banf,
        "excess": max(0.0, usable_target - real_needed_from_banf),
        "shortage_after_full_banf": max(0.0, net_real_requirement - usable_target),
        "post_banf_demand": -sum(
            r.qty for r in movements
            if r.row_no > target.row_no and r.qty < 0 and r.element not in SAFETY_TYPES
        ),
        "safety_stock": safety_stock,
        "stock": opening_stock,
        "other_receipts": other_receipts,
        "total_demand_excl_safety": real_demand,
    }


def fmt(x: float) -> str:
    return f"{x:,.3f}".rstrip("0").rstrip(".")


def main() -> None:
    p = argparse.ArgumentParser(description="MD04: benoetigten BANF-Anteil bis Stichtag berechnen")
    p.add_argument("excel", nargs="?", help="Pfad zur MD04-Excel-Datei")
    p.add_argument("--stichtag", help="z.B. 31.12.2026 oder 2026-12-31")
    p.add_argument("--banf", help="BANF/Position, z.B. 0017638725/00010; ohne Angabe: letzte BANF bis Stichtag")
    p.add_argument("--sheet", help="Tabellenblatt; Standard: erstes Blatt")
    args = p.parse_args()

    excel = Path(args.excel or input("Pfad zur MD04-Excel-Datei: ").strip().strip('"'))
    cutoff = parse_date(args.stichtag or input("Stichtag (TT.MM.JJJJ): ").strip())
    banf_ref = args.banf
    if banf_ref is None:
        entered = input("BANF/Position (Enter = letzte BANF bis Stichtag): ").strip()
        banf_ref = entered or None

    rows = read_md04(excel, args.sheet)
    target = choose_banf(rows, cutoff, banf_ref)
    r = calculate(rows, cutoff, target)

    print("\n=== Ergebnis ===")
    print(f"BANF:                         {target.reference}")
    print(f"BANF-Termin:                  {target.plan_date:%d.%m.%Y}")
    print(f"BANF-Menge:                   {fmt(target.qty)}")
    print(f"Stichtag:                     {cutoff:%d.%m.%Y}")
    print(f"Saldo ohne Ziel-BANF:         {fmt(r['balance_without_target'])}")
    print(f"Realer BANF-Bedarf:           {fmt(r['real_needed'])}")
    print(f"Ueberdeckung der BANF:        {fmt(r['excess'])}")
    print(f"Fehlmenge trotz voller BANF:  {fmt(r['shortage_after_full_banf'])}")
    print("\nKontrollwerte:")
    print(f"  Anfangsbestand:             {fmt(r['stock'])}")
    print(f"  Sonstige Zugaenge:          {fmt(r['other_receipts'])}")
    print(f"  Realer Bedarf:              {fmt(r['total_demand_excl_safety'])}")
    print(f"  Sicherheitsbestand (Info):  {fmt(r['safety_stock'])}")
    print(f"  Bruttobedarf nach BANF:      {fmt(r['post_banf_demand'])}")


if __name__ == "__main__":
    main()
